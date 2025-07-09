import argparse
import openpyxl
from datetime import datetime
from dateutil.relativedelta import relativedelta
from ames.harvesters import get_author_records
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

parser = argparse.ArgumentParser(
    prog="authors_funding",
    description="Harvests funding from CaltechAUTHORS",
)
parser.add_argument("author_identifier", type=str, help="The author identifier")

args = parser.parse_args()

author_identifier = args.author_identifier

# Get the current date
current_date = datetime.now()

# Need publications within the last 12 months
#start_date = current_date - relativedelta(months=12)
#start_date = start_date.strftime("%Y-%m-%d")
start_date = "2019-01-01"  # For testing purposes, use a fixed date

print(f"Searching for records after {start_date} for author {author_identifier}")

records = get_author_records(
    author_identifier=author_identifier, date=start_date, all_metadata=True
)


def update_funding(funder, new_funding_info, year, record_id):
    funder["record_ids"].append(record_id)
    # update year
    if funder["year"] < year:
        funder["year"] = year
    # add award
    if "award" in new_funding_info:
        number = new_funding_info["award"]["number"]
        if number not in funder["awards"]:
            funder["awards"][number] = [record_id]
        else:
            funder["awards"][number].append(record_id)


def create_funding(funding, year, record_id):
    if "award" in funding:
        awards = {funding["award"]["number"]: [record_id]}
    else:
        awards = {}
    if "id" in funding["funder"]:
        funder_id = funding["funder"]["id"]
    else:
        funder_id = ""
    return {
        "name": funding["funder"]["name"],
        "id": funder_id,
        "awards": awards,
        "year": year,
        "record_ids": [record_id],
    }


funders = {}
titles = {}

for article in records:
    year = article["metadata"]["publication_date"].split("-")[0]
    record_id = article["id"]
    titles[record_id] = article["metadata"].get("title", "No Title Provided")

    if "funding" in article["metadata"]:
        funding = article["metadata"]["funding"]
        for fund in funding:
            name = fund["funder"]["name"]
            if "id" in fund["funder"]:
                ror = fund["funder"]["id"]
                if ror in funders:
                    funder = funders[ror]
                    update_funding(funder, fund, year, record_id)
                else:
                    funders[ror] = create_funding(fund, year, record_id)
            else:
                if name in funders:
                    funder = funders[name]
                    update_funding(funder, fund, year, record_id)
                else:
                    funders[name] = create_funding(fund, year, record_id)

# Headers for the NSF collaborator report
header = [
    "Funder Name",
    "Funder ROR",
    "Grant Number",
    "Last Active",
    "Article IDs",
    "Article Titles",
]


data = []
# Add funder to the data list
for funder in funders:
    funder_data = funders[funder]
    awards = funder_data.get("awards")
    if awards:
        for award_number in awards.keys():
            record_str = ""
            title_str = ""
            for record_id in awards[award_number]:
                if record_str == "":
                    record_str = f"{record_id}"
                else:
                    record_str += f", {record_id}"
                if title_str == "":
                    title_str = titles[record_id]
                else:
                    title_str += f", {titles[record_id]}"
            data.append(
                [
                    funder_data["name"],
                    funder_data["id"],
                    award_number,
                    funder_data["year"],
                    record_str,
                    title_str,
                ]
            )
    else:
        record_str = ""
        title_str = ""
        for record_id in funder_data["record_ids"]:
            if record_str == "":
                record_str = f"{record_id}"
            else:
                record_str += f", {record_id}"
            if title_str == "":
                title_str = titles[record_id]
            else:
                title_str += f", {titles[record_id]}"
        data.append(
            [
                funder_data["name"],
                funder_data["id"],
                "",
                funder_data["year"],
                record_str,
                title_str,
            ]
        )

sorted_rows = sorted(data, key=lambda x: x[0])

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Funder Data"

# Populate the sheet with data
for row_index, row_data in enumerate([header] + sorted_rows, start=1):
    for col_index, cell_value in enumerate(row_data, start=1):
        cell = sheet.cell(row=row_index, column=col_index, value=cell_value)

        # Apply bold font to header row
        if row_index == 1:
            cell.font = Font(bold=True)

# Adjust column widths for better readability
for col_index in range(1, len(data[0]) + 1):
    column_letter = get_column_letter(col_index)
    sheet.column_dimensions[column_letter].width = 20

# Save the workbook to a file
output_filename = "funder_report.xlsx"
workbook.save(output_filename)

print(f"Report saved to {output_filename}")
