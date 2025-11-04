import argparse
import openpyxl
from unidecode import unidecode
from datetime import datetime
from dateutil.relativedelta import relativedelta
from ames.harvesters import get_author_records
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

parser = argparse.ArgumentParser(
    prog="nsf_c_s",
    description="Harvests content from CaltechAUTHORS for a NSF formetted collaborator report",
)
parser.add_argument("author_identifier", type=str, help="The author identifier")
parser.add_argument("--record_ids", action=argparse.BooleanOptionalAction)

args = parser.parse_args()
author_identifier = args.author_identifier
record_ids = args.record_ids

# Get the current date
current_date = datetime.now()

# Need publications within the last 48 months
start_date = current_date - relativedelta(months=48)
start_date = start_date.strftime("%Y-%m-%d")

print(f"Searching for records after {start_date} for author {author_identifier}")

records = get_author_records(
    author_identifier=author_identifier, date=start_date, all_metadata=True
)


def update_coauthor(coauthor, new_author_info, year, record_id):
    coauthor["record_ids"].append(record_id)
    # update year
    if coauthor["year"] < year:
        coauthor["year"] = year
    # add affiliation
    if "affiiations" in new_author_info:
        if new_author_info["affiliations"] not in coauthor["affiliations"]:
            coauthor["affiliations"].append(new_author_info["affiliations"])


def create_coauthor(author, year, record_id):
    if "affiliations" in author:
        affiliations = author["affiliations"]
    else:
        affiliations = []
    return {
        "name": author["person_or_org"]["name"],
        "affiliations": affiliations,
        "year": year,
        "record_ids": [record_id],
    }


coauthors = {}


for article in records:
    year = article["metadata"]["publication_date"].split("-")[0]
    authors = article["metadata"]["creators"]
    record_id = article["id"]
    for author in authors:
        name = author["person_or_org"]["name"]
        if author["person_or_org"]["type"] == "personal":
            if "identifiers" in author["person_or_org"]:
                identifiers = author["person_or_org"]["identifiers"]
                clpid = None
                orcid = None
                for identifier in identifiers:
                    if identifier["scheme"] == "clpid":
                        clpid = identifier["identifier"]
                    if identifier["scheme"] == "orcid":
                        orcid = identifier["identifier"]
                if clpid:
                    if clpid in coauthors:
                        coauthor = coauthors[clpid]
                        update_coauthor(coauthor, author, year, record_id)
                    elif orcid:
                        if orcid in coauthors:
                            # If the orcid record got created first, use that
                            coauthor = coauthors[orcid]
                            update_coauthor(coauthor, author, year, record_id)
                        else:
                            coauthors[clpid] = create_coauthor(author, year, record_id)
                    else:
                        coauthors[clpid] = create_coauthor(author, year, record_id)
                elif orcid:
                    if orcid in coauthors:
                        coauthor = coauthors[author_identifier]
                        update_coauthor(coauthor, author, year, record_id)
                    else:
                        coauthors[orcid] = create_coauthor(author, year, record_id)
            else:
                if name in coauthors:
                    coauthor = coauthors[name]
                    update_coauthor(coauthor, author, year, record_id)
                else:
                    coauthors[name] = create_coauthor(author, year, record_id)

# The author shouldn't be a coauthor
coauthors.pop(author_identifier)

# Headers for the NSF collaborator report
header = [
    "4",
    "Name:",
    "Organizational Affiliation",
    "Optional (email, Department",
    "Last Active",
]

if record_ids:
    header.append("CaltechAUTHORS Record IDs (do not include in NSF report)")

data = []
# Add coauthors to the data list
for coauthor in coauthors:
    a_string = ""
    for affiliation in coauthors[coauthor]["affiliations"]:
        if a_string == "":
            a_string = affiliation["name"]
        else:
            a_string += f", {affiliation['name']} "
    if record_ids:
        record_id_string = ""
        for record_id in coauthors[coauthor]["record_ids"]:
            if record_id_string == "":
                record_id_string = record_id
            else:
                record_id_string += f"; {record_id}"
        data.append(
            [
                "A:",
                unidecode(coauthors[coauthor]["name"]),
                unidecode(a_string),
                "",
                coauthors[coauthor]["year"],
                record_id_string,
            ]
        )
    else:
        data.append(
            [
                "A:",
                coauthors[coauthor]["name"],
                a_string,
                "",
                coauthors[coauthor]["year"],
            ]
        )

sorted_rows = sorted(data, key=lambda x: x[1])

# Create a new Excel workbook
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "NSF Collaborator Report Table 4"

# Populate the sheet with data
for row_index, row_data in enumerate([header] + sorted_rows, start=1):
    for col_index, cell_value in enumerate(row_data, start=1):
        cell = sheet.cell(row=row_index, column=col_index, value=cell_value)

        # Apply bold font to header row
        if row_index == 1:
            cell.font = Font(bold=True)

# Adjust column widths for better readability
for col_index in range(1, len(data[0]) + 1):
    column_letter = get_column_letter(col_index)
    sheet.column_dimensions[column_letter].width = 20

# Save the workbook to a file
output_filename = f"{author_identifier}_nsf_collaborator_report.xlsx"
workbook.save(output_filename)

print(f"Report saved to {output_filename}")
